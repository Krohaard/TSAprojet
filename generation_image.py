import os
import fnmatch
import random
import requests
from urllib import request
import json
import time
from datetime import datetime
from openpyxl import load_workbook
import shutil
from pathlib import Path

def chercher_dossier_relatif(nom_partiel: str) -> list[Path]:
    """
    Cherche récursivement des dossiers dont le nom contient nom_partiel,
    depuis le dossier courant, et retourne les chemins relatifs.

    :param nom_partiel: str -> texte à chercher dans le nom du dossier
    :return: list[Path] -> chemins relatifs des dossiers trouvés
    :raises FileNotFoundError: si aucun dossier trouvé
    """
    cwd = Path.cwd()
    resultats = [p.relative_to(cwd) for p in cwd.rglob("*") if p.is_dir() and nom_partiel in p.name]

    if not resultats:
        raise FileNotFoundError(f"Aucun dossier contenant '{nom_partiel}' trouvé dans le dossier courant")

    return resultats

def trouver_diff_relatif(nom: str) -> Path:
    """
    Recherche récursivement un fichier ou dossier par son NOM
    depuis le dossier courant et retourne son chemin relatif.

    :param nom: nom du fichier ou dossier à chercher (str)
    :return: Path (chemin relatif)
    :raises FileNotFoundError: si non trouvé
    :raises ValueError: si plusieurs résultats trouvés
    """
    cwd = Path.cwd()

    # Recherche récursive
    resultats = list(cwd.rglob(nom))

    if not resultats:
        raise FileNotFoundError(f"'{nom}' introuvable dans le dossier courant")

    if len(resultats) > 1:
        raise ValueError(
            f"Plusieurs éléments nommés '{nom}' trouvés :\n" +
            "\n".join(str(p.relative_to(cwd)) for p in resultats)
        )

    # Retourne le chemin relatif
    return resultats[0].relative_to(cwd)



def deplacer_fichier(source: str, destination: str) -> None:
    try:
        shutil.move(source,destination)
    except Exception as e:
        print(f"une erreur est survenue : {e}")

def chercher_fichier(dossier: str, nom_partiel: str) -> list[str]:
    """
    Cherche récursivement un fichier dans un dossier Windows par nom partiel.
    
    dossier : str -> chemin du dossier de départ
    nom_partiel : str -> texte à chercher dans le nom du fichier
    
    Retourne : liste de chemins complets des fichiers correspondants
    """
    fichiers_trouves: list[str] = []
    
    for root, dirs, files in os.walk(dossier):
        for file in files:
            if fnmatch.fnmatch(file, f"*{nom_partiel}*"):
                fichiers_trouves.append(os.path.join(root, file))
    
    return fichiers_trouves


def extract_workflow_info(workflow: str) -> list:
    objets: list = []
    try:
        with open(workflow, "r", encoding="utf-8") as f:
            data = json.load(f)
        objets.append( data.get("id") )
        objets.append( data.get("revision") )
        objets.append( data.get("last_node_id") )
        objets.append( data.get("last_link_id") )
        objets.append( data.get("version") )
        return objets
    except FileNotFoundError:
        print(f"Erreur : le fichier '{workflow}' n'a pas été trouvé.")
        return []
    except json.JSONDecodeError as e:
        print(f"Erreur de lecture JSON : {e}")
        return []

def find_node_ids(data, node_type=None, title_contains=None)-> int:
    results = []
    ksearch = ''
    knode:int
    if node_type==None and title_contains!=None:
        ksearch = 'title'
    if node_type!=None and title_contains==None:
        ksearch = 'type'
    if node_type!=None and title_contains!=None:
        ksearch = 'both'
    if ksearch!='':
        for knode in data:
            if ksearch=='title':
                if data[knode]['_meta']['title'] is not None and title_contains.lower()!=data[knode]['_meta']['title'].lower() :
                    continue
            elif ksearch=='type':
                if data[knode]['class_type'] is not None and data[knode]['class_type'].lower()!=node_type.lower() :
                    continue
            elif ksearch=='both':
                if data[knode]['_meta']['title'] is not None and data[knode]['class_type'] is not None and data[knode]['class_type'].lower()!=node_type.lower() and title_contains.lower()!=data[knode]['_meta']['title'].lower():
                    continue
            if data[knode]['_meta']['title'] is not None and data[knode]['class_type'] is not None:
                results.append(knode)
    return results[0]

def update_widgets_values(data, node_id: int, updates: dict) -> bool:
    """
    Modifie un ou plusieurs éléments de widgets_values pour un node donné.

    data : dict (JSON chargé)
    node_id : int
    updates : dict {index: nouvelle_valeur}

    Retourne : True si modifié, False sinon
    """
    knode:int
    index:str
    new_value = None

    for knode in data:
        if knode == node_id:
            for index, new_value in updates.items():
                data[knode]['inputs'][index] = new_value

            return True

    return False

def save_json(data, file_path:str) -> None:
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(
            data,
            f,
            indent=2,              # lisible
            ensure_ascii=False     # garde les accents
        )

def open_json(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)

def generate_image_from_workflow(file_path:str, prompt_description:str, output_dir=r'.\ComfyUI_windows_portable\ComfyUI\output', model=0,seed_input=random.randint(123456789, 23654987321)) -> tuple[str,str]:
    output_file:str = ''
    input_model:str = ''
    list_model = ['sd_xl_base_1.0.safetensors','Juggernaut-XL_v9_RunDiffusionPhoto_v2.safetensors','awpainting_v14.safetensors']
    kworkflow = open_json(file_path)
#    update_widgets_values(kworkflow, find_node_ids(kworkflow, title_contains='seed_image'), {'value': seed_input})
#    update_widgets_values(kworkflow, find_node_ids(kworkflow, title_contains='Inputmodel'), {'ckpt_name': list_model[model]})
#    input_model = kworkflow[find_node_ids(kworkflow, title_contains='Inputmodel')]['inputs']['ckpt_name']
    update_widgets_values(kworkflow, find_node_ids(kworkflow, title_contains='InputPrompt'), {'value': prompt_description})
    kfilename = datetime.now().strftime("%Y%m%d%H%M%S") #+ "_" + str(kworkflow[find_node_ids(kworkflow, title_contains='seed_image')]['inputs']['value'])
    update_widgets_values(kworkflow, find_node_ids(kworkflow, title_contains='InputStringName'), {'value': kfilename})
    #save_json(kworkflow, file_path)

    COMFYUI_QUEUE_URL = "http://127.0.0.1:8188"
    payload = {
        "prompt": kworkflow
    }
    data = json.dumps(payload).encode("utf-8")

    # Envoyer le workflow à la queue
    req = request.Request(
        f"{COMFYUI_QUEUE_URL}/prompt",
        data=data,
        headers={"Content-Type": "application/json"}
    )

    with request.urlopen(req) as res:
        response_text = res.read().decode("utf-8")

    print("Réponse ComfyUI:", response_text)

    prompt_id = json.loads(response_text)['prompt_id']

    while True:
        r = requests.get(f"http://127.0.0.1:8188/history/{prompt_id}")
        r.raise_for_status()
        history = r.json()

        if prompt_id in history:
            break

        time.sleep(0.5)
    output_file = chercher_fichier(output_dir, kfilename)[0]
    return output_file,input_model

def kmain():
    fileAPI:str = ''
    kinput_model: str = ''
    output_dir_image: str = ''
    output_image: str = ''
#    kseed_all:list[int] = [random.randint(123456789, 23654987321) for _ in range(4)]
    kseed_all:list[int] = [random.randint(123456789, 23654987321) for _ in range(1)]
    kfile_path:list[str]= []
    liste_dossiers:list[str] = []
#    kfile_path.append( str(trouver_diff_relatif(r'.\ImagePersoAPI.json')))
#    kfile_path.append( str(trouver_diff_relatif(r'.\ImagePersoAPIWithoutLayerMask.json')))
#    kfile_path.append( str(trouver_diff_relatif(r'.\ImagePersoAPIwithoutLora.json')))
#    kfile_path.append( str(trouver_diff_relatif(r'.\ImagePersoAPIwithoutLoraWithoutLayerMask.json')))
    kfile_path.append( r'C:\Users\lione\Downloads\ImagePersoTest.json')
    kExcel_path = str(trouver_diff_relatif(r'.\bdd_images.xlsx'))
    output_dir_image = str(trouver_diff_relatif(r'.\ComfyUI_windows_portable\ComfyUI\output'))
    wb = load_workbook(kExcel_path)
    ws = wb.active
#    liste_dossiers.append(str(trouver_diff_relatif(chercher_dossier_relatif("0" + str(0+3) + "-")[0].name)))
#    liste_dossiers.append(str(trouver_diff_relatif(chercher_dossier_relatif("0" + str(1+3) + "-")[0].name)))
    kkstart = 9
    for fileAPI in kfile_path:
        print(fileAPI)
        k=0
        # Lire toutes les cellules non vides de la colonne E
        for cell in ws["D"]:
            i=0
            print(cell.value)
            for kseed in kseed_all:
#            for kseed in kseed_all:
                print(kseed)
#                for kk in range(2):
                for kk in range(1):
                    koutput_path = r'C:\Users\lione\Documents\TSAproject\image\06-Test'
                    print(koutput_path)
                    if cell.value is not None and k>0 and ws.cell(row=cell.row,column=5).value is not None and ws.cell(row=cell.row, column=kkstart+i+2).value is None and ws.cell(row=cell.row, column=6).value is None:
                        output_image,kinput_model = generate_image_from_workflow(file_path=fileAPI, prompt_description=ws.cell(row=cell.row,column=5).value, output_dir=output_dir_image,model = kk,seed_input=kseed)
                        deplacer_fichier(output_image, koutput_path + '\\' + os.path.basename(output_image))
                        ws.cell(row=cell.row, column=kkstart+i).value = kseed
                        ws.cell(row=cell.row, column=kkstart+i+1).value = kinput_model
                        ws.cell(row=cell.row, column=kkstart+i+2).value = koutput_path + '\\' + os.path.basename(output_image)
                        print(f"Image générée et sauvegardée dans : {koutput_path + os.path.basename(output_image)}")
                        wb.save(kExcel_path)
                    i=i+3
#            if k>=6:
#                break
            k=k+1

if __name__ == "__main__":
    kmain()
    print("Fin du programme.")